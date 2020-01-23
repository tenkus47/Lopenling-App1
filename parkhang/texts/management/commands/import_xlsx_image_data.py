import openpyxl
import csv
import re
import os

from django.core.management.base import BaseCommand

from texts.models import Text, Source, Witness, Annotation, AnnotationType

SOURCE_NAMES = {
    'D': 'སྡེ་དགེ',
    'Q': 'པེ་ཅིན',
    'C': 'ཅོ་ནེ',
    'N': 'སྣར་ཐང'
}

class Command(BaseCommand):
    IMAGE_START_PRE_KEY = "bdrcimg_pre"
    IMAGE_START_NUMBER_KEY = "bdrcimg_number"
    IMAGE_START_SUFFIX_KEY = "bdrcimg_suffix"

    def add_arguments(self, parser):
        parser.add_argument('source_dir', nargs='?')

    def handle(self, *args, **options):
        source_dir = options['source_dir']
        sources = {}
        for source_code in SOURCE_NAMES:
            source_name = SOURCE_NAMES[source_code]
            try:
                source = Source.objects.get(name=source_name)
                sources[source_code] = source
            except:
                continue
        data_files = next(os.walk(source_dir))[2]
        processed_texts = 0
        for data_filename in data_files:
            if not re.match(r'[A-Z][0-9]+.*\.xlsx', data_filename):
                continue
            text_code = re.search(r'([A-Z][0-9]+)', data_filename).group(0).lower()
            try:
                text = Text.objects.get(code=text_code)
            except:
                print(f"No text found for code {text_code}")
                continue

            processed_texts += 1
            data_filepath = os.path.join(source_dir, data_filename)
            wb = openpyxl.load_workbook(filename=data_filepath)
            sheetnames = wb.get_sheet_names()
            
            for sheetname in sheetnames:
                if sheetname not in SOURCE_NAMES or sheetname not in sources:
                    continue
                source = sources[sheetname]
                sheet = wb[sheetname]
                rows = list(sheet.rows)
                if rows[0]:
                    row = rows[0]
                    cell_values = [cell.value for cell in row]
                    if len(cell_values) < 3 and not re.match(r'http', cell_values[2]):
                        continue
                    first_image = cell_values[2]
                    try:
                        witness = Witness.objects.get(text=text, source=source)
                        properties = {}
                        if witness.properties:
                            properties = witness.properties
                        matches = re.findall("(bdr:[^_]+_[^:]+\:\:.*?)([0-9]+)\.([a-z]{3})", first_image)
                        if (len(matches) is 1):
                            match_groups = matches[0]
                            properties[self.IMAGE_START_PRE_KEY] = match_groups[0]
                            properties[self.IMAGE_START_NUMBER_KEY] = match_groups[1]
                            properties[self.IMAGE_START_SUFFIX_KEY] = match_groups[2]
                        witness.properties = properties
                        witness.save()
                    except:
                        continue

        print(f"Processed {processed_texts} texts")